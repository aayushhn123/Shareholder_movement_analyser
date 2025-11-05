import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
import matplotlib.pyplot as plt
import io
import os
import tempfile
from fpdf import FPDF
from datetime import datetime

def parse_sheet_date(sheet_name):
    """Parse different date formats from sheet names"""
    formats = [
        '%d_%B_%Y',      # 15_August_2025
        '%d-%b-%Y',      # 12-Sep-2025
        '%d-%B-%Y',      # 12-September-2025
        '%d_%b_%Y',      # 15_Sep_2025
        '%Y-%m-%d',      # 2025-09-12
        '%m-%d-%Y',      # 09-12-2025
        '%d/%m/%Y',      # 12/09/2025
        '%m/%d/%Y',      # 09/12/2025
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(sheet_name, fmt)
        except ValueError:
            continue
    
    return None

def analyze_shareholder_changes(excel_file):
    try:
        xls = pd.ExcelFile(excel_file)
        sheet_names = xls.sheet_names
        
        # Parse and sort dates properly
        sheet_dates = []
        for sheet_name in sheet_names:
            parsed_date = parse_sheet_date(sheet_name)
            if parsed_date:
                sheet_dates.append((sheet_name, parsed_date))
            else:
                st.warning(f"Could not parse date from sheet name: '{sheet_name}'. Skipping this sheet.")
        
        if len(sheet_dates) < 3:
            raise ValueError("At least three sheets with valid dates are required for consecutive and overall comparisons.")
        
        # Sort by parsed date
        sheet_dates.sort(key=lambda x: x[1])
        dates = [sheet_name for sheet_name, _ in sheet_dates]
        
        st.info(f"Processing sheets in chronological order: {dates}")
        
        df_dict = {date: pd.read_excel(xls, sheet_name=date) for date in dates}
        
        id_col = 'PAN NO'
        name_col = 'NAME'
        holding_col = 'HOLDING '
        
        for date, df in df_dict.items():
            if id_col not in df.columns or name_col not in df.columns or holding_col not in df.columns:
                raise ValueError(f"Sheet '{date}' is missing one or more required columns: {id_col}, {name_col}, {holding_col}")
        
        all_data = []
        for date in dates:
            df = df_dict[date].copy()
            df = df.groupby([id_col, name_col], as_index=False)[holding_col].sum()
            df['Date'] = date
            all_data.append(df)
        
        all_df = pd.concat(all_data, ignore_index=True)
        
        # Create a mapping of PAN NO to NAME
        names_dict = {}
        for _, row in all_df[[id_col, name_col]].drop_duplicates().iterrows():
            names_dict[row[id_col]] = row[name_col]
        
        changes = []
        
        # ALL Consecutive comparisons
        for i in range(len(dates) - 1):
            date1 = dates[i]
            date2 = dates[i + 1]
            date_label = f"{date1} to {date2}"
            
            # Create proper dictionaries from grouped data
            h1_df = df_dict[date1].groupby(id_col, as_index=False)[holding_col].sum()
            h1_dict = dict(zip(h1_df[id_col], h1_df[holding_col]))
            
            h2_df = df_dict[date2].groupby(id_col, as_index=False)[holding_col].sum()
            h2_dict = dict(zip(h2_df[id_col], h2_df[holding_col]))
            
            all_ids = set(h1_dict.keys()) | set(h2_dict.keys())
            
            for id_ in all_ids:
                hold1 = h1_dict.get(id_, 0)
                hold2 = h2_dict.get(id_, 0)
                
                if hold1 == hold2:
                    continue
                
                if hold1 == 0 and hold2 > 0:
                    action = 'entry'
                    change_amount = hold2
                elif hold2 > hold1:
                    action = 'increase'
                    change_amount = hold2 - hold1
                elif hold2 < hold1:
                    action = 'exit' if hold2 == 0 else 'decrease'
                    change_amount = hold2 - hold1
                
                name = names_dict.get(id_, 'Unknown')
                
                changes.append({
                    'Name': name,
                    'Action': action,
                    'Change Amount': change_amount,
                    'Date Transition': date_label
                })
        
        # Overall comparison (first to last)
        date1 = dates[0]
        date2 = dates[-1]
        date_label = f"Overall_{date1} to {date2}"
        
        h1_df = df_dict[date1].groupby(id_col, as_index=False)[holding_col].sum()
        h1_dict = dict(zip(h1_df[id_col], h1_df[holding_col]))
        
        h2_df = df_dict[date2].groupby(id_col, as_index=False)[holding_col].sum()
        h2_dict = dict(zip(h2_df[id_col], h2_df[holding_col]))
        
        all_ids = set(h1_dict.keys()) | set(h2_dict.keys())
        
        for id_ in all_ids:
            hold1 = h1_dict.get(id_, 0)
            hold2 = h2_dict.get(id_, 0)
            
            if hold1 == hold2:
                continue
            
            if hold1 == 0 and hold2 > 0:
                action = 'entry'
                change_amount = hold2
            elif hold2 > hold1:
                action = 'increase'
                change_amount = hold2 - hold1
            elif hold2 < hold1:
                action = 'exit' if hold2 == 0 else 'decrease'
                change_amount = hold2 - hold1
            
            name = names_dict.get(id_, 'Unknown')
            
            changes.append({
                'Name': name,
                'Action': action,
                'Change Amount': change_amount,
                'Date Transition': date_label
            })
        
        changes_df = pd.DataFrame(changes)
        
        if changes_df.empty:
            return None, None, None, None, None, None, None, "No changes detected."
        
        # Pivot table for output
        pivot_df = changes_df.pivot_table(
            index=['Name', 'Action'],
            columns='Date Transition',
            values='Change Amount',
            aggfunc='first'
        ).reset_index()
        pivot_df = pivot_df.fillna(0)
        
        # Reorder columns
        date_columns = [col for col in pivot_df.columns if col not in ['Name', 'Action']]
        consecutive_labels = [f"{dates[i]} to {dates[i+1]}" for i in range(len(dates)-1)]
        overall_labels = [col for col in date_columns if col.startswith('Overall_')]
        
        ordered_columns = ['Name', 'Action'] + consecutive_labels + overall_labels
        pivot_df = pivot_df[ordered_columns]
        
        # Aggregate counts for visualization
        counts = changes_df.groupby(['Date Transition', 'Action']).size().unstack(fill_value=0)
        
        date_pairs = consecutive_labels + overall_labels
        date_pairs = [dp for dp in date_pairs if dp in counts.index]
        
        increases = counts.get('increase', pd.Series(0, index=counts.index)).reindex(date_pairs, fill_value=0).tolist()
        decreases = counts.get('decrease', pd.Series(0, index=counts.index)).reindex(date_pairs, fill_value=0).tolist()
        exits = counts.get('exit', pd.Series(0, index=counts.index)).reindex(date_pairs, fill_value=0).tolist()
        entries = counts.get('entry', pd.Series(0, index=counts.index)).reindex(date_pairs, fill_value=0).tolist()
        
        # Create Plotly chart
        fig = go.Figure()
        
        x = date_pairs
        width = 0.2
        
        fig.add_trace(go.Bar(
            x=x, y=increases, name='Increase', marker_color='#4CAF50',
            offset=-1.5*width, width=width,
            text=[int(i) if i > 0 else '' for i in increases],
            textposition='auto',
            hovertemplate='%{y} shareholders<br>Date: %{x}<extra>Increase</extra>'
        ))
        fig.add_trace(go.Bar(
            x=x, y=decreases, name='Decrease', marker_color='#F44336',
            offset=-0.5*width, width=width,
            text=[int(i) if i > 0 else '' for i in decreases],
            textposition='auto',
            hovertemplate='%{y} shareholders<br>Date: %{x}<extra>Decrease</extra>'
        ))
        fig.add_trace(go.Bar(
            x=x, y=exits, name='Exit', marker_color='#2196F3',
            offset=0.5*width, width=width,
            text=[int(i) if i > 0 else '' for i in exits],
            textposition='auto',
            hovertemplate='%{y} shareholders<br>Date: %{x}<extra>Exit</extra>'
        ))
        fig.add_trace(go.Bar(
            x=x, y=entries, name='Entry', marker_color='#FFC107',
            offset=1.5*width, width=width,
            text=[int(i) if i > 0 else '' for i in entries],
            textposition='auto',
            hovertemplate='%{y} shareholders<br>Date: %{x}<extra>Entry</extra>'
        ))
        
        fig.update_layout(
            title='Shareholder Changes Across Dates',
            xaxis_title='Date Transition',
            yaxis_title='Number of Shareholders',
            barmode='group',
            xaxis_tickangle=-45,
            legend=dict(x=0, y=1.0, bgcolor='rgba(255, 255, 255, 0)', bordercolor='rgba(255, 255, 255, 0)'),
            margin=dict(b=150),
            showlegend=True,
            plot_bgcolor='white',
            font=dict(family="Arial", size=12)
        )
        
        return pivot_df, fig, None, increases, decreases, exits, entries, date_pairs
    
    except ValueError as e:
        return None, None, None, None, None, None, None, f"Error: {str(e)}"
    except Exception as e:
        return None, None, None, None, None, None, None, f"An unexpected error occurred: {str(e)}"

def generate_matplotlib_plot(increases, decreases, exits, entries, date_pairs, output):
    try:
        plt.figure(figsize=(12, 8))
        x = np.arange(len(date_pairs))
        width = 0.2
        
        bars1 = plt.bar(x - 1.5*width, increases, width, label='Increase', color='#4CAF50')
        bars2 = plt.bar(x - 0.5*width, decreases, width, label='Decrease', color='#F44336')
        bars3 = plt.bar(x + 0.5*width, exits, width, label='Exit', color='#2196F3')
        bars4 = plt.bar(x + 1.5*width, entries, width, label='Entry', color='#FFC107')
        
        plt.xlabel('Date Transition')
        plt.ylabel('Number of Shareholders')
        plt.title('Shareholder Changes Across Dates')
        plt.xticks(x, date_pairs, rotation=45, ha='right')
        plt.legend()
        plt.tight_layout()
        
        for bars, heights in [(bars1, increases), (bars2, decreases), (bars3, exits), (bars4, entries)]:
            for bar, height in zip(bars, heights):
                if height > 0:
                    plt.text(bar.get_x() + bar.get_width() / 2, height,
                             f'{int(height)}', ha='center', va='bottom')
        
        plt.savefig(output, format='png', dpi=300, bbox_inches='tight')
        plt.close()
        return True, None
    except Exception as e:
        return False, f"Error generating matplotlib plot: {str(e)}"

def generate_pdf(pivot_df, fig, increases, decreases, exits, entries, date_pairs):
    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            temp_excel = os.path.join(tmpdirname, "temp_changes.xlsx")
            with pd.ExcelWriter(temp_excel, engine='openpyxl') as writer:
                pivot_df.to_excel(writer, index=False, sheet_name='Changes')
            
            plot_path = os.path.join(tmpdirname, "plot.png")
            success, error = generate_matplotlib_plot(increases, decreases, exits, entries, date_pairs, plot_path)
            if not success:
                return None, error
            
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Shareholder Changes Report", ln=True, align='C')
            pdf.set_font("Arial", "", 12)
            pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%d-%B-%Y')}", ln=True, align='C')
            pdf.ln(10)
            
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Legend", ln=True, align='L')
            pdf.set_font("Arial", "", 10)
            
            legend_data = [
                ['Value/Color', 'Meaning'],
                ['0', 'No change in Holding'],
                ['Green text', 'Increase in holding'],
                ['Red text', 'Decrease in holding'],
                ['Yellow text (entry)', 'New shareholder entry'],
                ['Blue text (exit)', 'Shareholder exit']
            ]
            
            legend_col_widths = [60, 100]
            
            pdf.set_fill_color(200, 200, 200)
            for text, width in zip(legend_data[0], legend_col_widths):
                pdf.cell(width, 10, text, border=1, align='L', fill=True)
            pdf.ln()
            
            legend_colors = [
                (0, 0, 0), (76, 175, 80), (244, 67, 54), (255, 193, 7), (33, 150, 243)
            ]
            
            for i, (row_data, color) in enumerate(zip(legend_data[1:], legend_colors)):
                for j, (text, width) in enumerate(zip(row_data, legend_col_widths)):
                    if j == 0:
                        pdf.set_text_color(*color)
                    else:
                        pdf.set_text_color(0, 0, 0)
                    pdf.cell(width, 10, text, border=1, align='L')
                pdf.ln()
            
            pdf.set_text_color(0, 0, 0)
            pdf.ln(5)
            
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Summary Table", ln=True, align='L')
            pdf.set_font("Arial", "", 10)
            
            page_width = pdf.w - 2 * pdf.l_margin
            num_data_cols = len(pivot_df.columns) - 2
            col_widths = [page_width * 0.25, page_width * 0.15] + [page_width * 0.60 / num_data_cols] * num_data_cols
            
            pdf.set_fill_color(200, 200, 200)
            start_y = pdf.get_y()
            overall_col_idx = len(pivot_df.columns) - 1
            overall_text = str(pivot_df.columns[overall_col_idx])
            overall_width = col_widths[overall_col_idx]
            overall_text_width = pdf.get_string_width(overall_text)
            overall_num_lines = max(1, int(overall_text_width / overall_width) + 1)
            max_height = 5 * overall_num_lines
            
            current_x = pdf.get_x()
            for i, (col, width) in enumerate(zip(pivot_df.columns, col_widths)):
                text = str(col)
                pdf.rect(current_x, start_y, width, max_height, 'F')
                pdf.set_xy(current_x, start_y)
                pdf.set_text_color(0, 0, 0)
                pdf.multi_cell(width, 5, text, border=1, align='C', ln=0)
                current_x += width
            
            pdf.set_y(start_y + max_height)
            
            for _, row in pivot_df.iterrows():
                action = row['Action']
                
                for i, (val, width) in enumerate(zip(row, col_widths)):
                    val_str = str(val)
                    if len(val_str) > 30:
                        val_str = val_str[:27] + "..."
                    
                    if i == 1:
                        if action == 'entry':
                            pdf.set_text_color(255, 193, 7)
                        elif action == 'exit':
                            pdf.set_text_color(33, 150, 243)
                        elif action == 'increase':
                            pdf.set_text_color(76, 175, 80)
                        elif action == 'decrease':
                            pdf.set_text_color(244, 67, 54)
                        else:
                            pdf.set_text_color(0, 0, 0)
                    elif i >= 2:
                        val_numeric = row.iloc[i]
                        try:
                            is_non_zero = float(val_numeric) != 0.0
                        except (ValueError, TypeError):
                            is_non_zero = False
                        
                        if is_non_zero:
                            if action in ['increase', 'entry']:
                                pdf.set_text_color(76, 175, 80)
                            elif action in ['decrease', 'exit']:
                                pdf.set_text_color(244, 67, 54)
                            else:
                                pdf.set_text_color(0, 0, 0)
                        else:
                            pdf.set_text_color(0, 0, 0)
                    else:
                        pdf.set_text_color(0, 0, 0)
                    
                    pdf.cell(width, 10, val_str, border=1, align='L')
                
                pdf.ln()
            
            pdf.set_text_color(0, 0, 0)
            
            pdf.add_page()
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "Visualization", ln=True, align='L')
            
            pdf.image(plot_path, x=10, y=pdf.get_y(), w=page_width)
            
            pdf_path = os.path.join(tmpdirname, "report.pdf")
            pdf.output(pdf_path)
            with open(pdf_path, "rb") as f:
                pdf_data = f.read()
            return pdf_data, None
    
    except Exception as e:
        return None, f"Error generating PDF: {str(e)}"

# Streamlit App
st.set_page_config(page_title="Shareholder Changes Analyzer", layout="wide")

st.title("Shareholder Changes Analyzer")
st.markdown("""
Upload an Excel file with at least three sheets representing different dates. 
The sheets will be automatically sorted by date, and the app will analyze all consecutive transitions plus an overall comparison.

**Supported date formats in sheet names:**
- `15_August_2025` or `15_Sep_2025`  
- `12-Sep-2025` or `12-September-2025`
- `2025-09-12` or `09-12-2025`
- `12/09/2025` or `09/12/2025`

Click the 'Analyze' button to process the data and generate a summary table, downloadable Excel, interactive visualization, and a PDF report.
""")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

if 'files_generated' not in st.session_state:
    st.session_state.files_generated = False
    st.session_state.excel_data = None
    st.session_state.png_data = None
    st.session_state.pdf_data = None
    st.session_state.pivot_df = None
    st.session_state.fig = None

if uploaded_file is not None:
    if st.button("Analyze"):
        with st.spinner("Analyzing data..."):
            pivot_df, fig, error, increases, decreases, exits, entries, date_pairs = analyze_shareholder_changes(uploaded_file)
        
        if error:
            st.error(error)
        else:
            st.success("Analysis complete!")
            
            st.session_state.pivot_df = pivot_df
            st.session_state.fig = fig
            
            output = io.BytesIO()
            try:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pivot_df.to_excel(writer, index=False, sheet_name='Changes', startrow=8)
        
                    from openpyxl.styles import Font, Alignment
                    workbook = writer.book
                    worksheet = writer.sheets['Changes']
        
                    worksheet.sheet_state = 'visible'
        
                    worksheet.merge_cells('A1:B1')
                    legend_title = worksheet['A1']
                    legend_title.value = 'Legend'
                    legend_title.font = Font(bold=True, size=14)
                    legend_title.alignment = Alignment(horizontal='center')
        
                    worksheet['A2'] = 'Value/Color'
                    worksheet['B2'] = 'Meaning'
                    worksheet['A2'].font = Font(bold=True)
                    worksheet['B2'].font = Font(bold=True)
        
                    worksheet['A3'] = '0'
                    worksheet['B3'] = 'No change in Holding'
        
                    worksheet['A4'] = 'Green text'
                    worksheet['B4'] = 'Increase in holding'
                    worksheet['A4'].font = Font(color='4CAF50')
        
                    worksheet['A5'] = 'Red text'
                    worksheet['B5'] = 'Decrease in holding'
                    worksheet['A5'].font = Font(color='F44336')
        
                    worksheet['A6'] = 'Yellow text (entry)'
                    worksheet['B6'] = 'New shareholder entry'
                    worksheet['A6'].font = Font(color='FFC107')
        
                    worksheet['A7'] = 'Blue text (exit)'
                    worksheet['B7'] = 'Shareholder exit'
                    worksheet['A7'].font = Font(color='2196F3')
        
                    action_col = 2
                    data_start_col = 3
        
                    for row_idx in range(10, len(pivot_df) + 10):
                        action_cell = worksheet.cell(row=row_idx, column=action_col)
                        action_value = action_cell.value
            
                        if action_value == 'entry':
                            action_cell.font = Font(color='FFC107')
                        elif action_value == 'exit':
                            action_cell.font = Font(color='2196F3')
                        elif action_value == 'increase':
                            action_cell.font = Font(color='4CAF50')
                        elif action_value == 'decrease':
                            action_cell.font = Font(color='F44336')
            
                        for col_idx in range(data_start_col, len(pivot_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell_value = cell.value
                
                            if cell_value != 0 and cell_value != 0.0:
                                if action_value == 'entry':
                                    text_color = 'FFC107'
                                elif action_value == 'exit':
                                    text_color = '2196F3'
                                elif action_value in ['increase']:
                                    text_color = '4CAF50'
                                elif action_value in ['decrease']:
                                    text_color = 'F44336'
                                else:
                                    text_color = '000000'
                            else:
                                text_color = '000000'
                
                            cell.font = Font(color=text_color)
    
                    output.seek(0)
                    st.session_state.excel_data = output.getvalue()
    
            except Exception as e:
                st.error(f"Error generating Excel file: {str(e)}")
                st.session_state.excel_data = None
            
            plot_output = io.BytesIO()
            success, error = generate_matplotlib_plot(increases, decreases, exits, entries, date_pairs, plot_output)
            if success:
                plot_output.seek(0)
                st.session_state.png_data = plot_output.getvalue()
            else:
                st.error(error)
            
            with st.spinner("Generating PDF report..."):
                pdf_data, pdf_error = generate_pdf(pivot_df, fig, increases, decreases, exits, entries, date_pairs)
            if pdf_error:
                st.error(pdf_error)
            else:
                st.session_state.pdf_data = pdf_data
            
            st.session_state.files_generated = True
    
    if st.session_state.files_generated:
        st.header("Summary Table")
        st.dataframe(st.session_state.pivot_df, use_container_width=True)
        
        st.download_button(
            label="Download Changes Excel",
            data=st.session_state.excel_data,
            file_name="shareholder_changes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel"
        )
        
        st.header("Shareholder Changes Visualization")
        st.markdown("Hover over the bars to see the number of shareholders and date transitions.")
        st.plotly_chart(st.session_state.fig, use_container_width=True)
        
        if st.session_state.png_data:
            st.download_button(
                label="Download Plot PNG",
                data=st.session_state.png_data,
                file_name="shareholder_changes_plot.png",
                mime="image/png",
                key="download_plot"
            )
        
        if st.session_state.pdf_data:
            st.download_button(
                label="Download PDF Report",
                data=st.session_state.pdf_data,
                file_name="shareholder_changes_report.pdf",
                mime="application/pdf",
                key="download_pdf"
            )
else:
    st.info("Please upload an Excel file and click 'Analyze' to begin.")
